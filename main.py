
import streamlit as st
from iapws import IAPWS97
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from iapws import IAPWS97 as WSP
import math as M

from sympy import *

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import os
import time


def print_xlsx():
    newbook = openpyxl.Workbook()
    newbook.remove(newbook.active)

    sheet_1 = newbook.create_sheet("Вариант")
    for column in range(1, st.session_state.sheet.max_column + 1):
        sheet_1.cell(row=1, column=column).value = st.session_state.sheet.cell(row=1, column=column).value
        sheet_1.cell(row=2, column=column).value = st.session_state.sheet.cell(row=index_row, column=column).value

    sheet_2 = newbook.create_sheet("Задание 1")
    sheet_2['A1'] = 'Максимальный КПД'
    sheet_2['A2'] = 'Расход пара на входе в турбину (G0) при макс. КПД'
    sheet_2['A3'] = 'Расход пара на входе в конденсатор (Gк) при макс. КПД'
    sheet_2['A4'] = 'Давление пром. перегрева при макс. КПД'
    sheet_2['B1'] = st.session_state.eta_fmax
    sheet_2['B2'] = st.session_state.G0max
    sheet_2['B3'] = st.session_state.Gkmax
    sheet_2['B4'] = st.session_state.ppp_fmax
    sheet_2['C1'] = '%'
    sheet_2['C2'] = 'кг/с'
    sheet_2['C3'] = 'кг/с'
    sheet_2['C4'] = 'МПа'
    sheet_2['A7'] = 'Табл. Зависимость КПД от Pпп'
    for r in dataframe_to_rows(st.session_state.ppp_eta, index=False, header=True):
        sheet_2.append(r)
    img1 = Image('Зависимость КПД от давления пром. перегрева.png')
    img2 = Image('h - s диаграмма.png')
    sheet_2.add_image(img1, 'H1')
    sheet_2.add_image(img2, 'S1')

    sheet_3 = newbook.create_sheet("Задание 2")
    sheet_3['A1'] = 'Внутренний относительный КПД ступени eta_oi'
    sheet_3['A2'] = 'Внутреняя мощность ступени N_i '
    sheet_3['B1'] = (st.session_state.eta_oi * 100)
    sheet_3['B2'] = st.session_state.N_i
    sheet_3['C1'] = '%'
    sheet_3['C2'] = 'кВт'
    sheet_3['A5'] = 'Табл. Зависимость ηол от U/cф'
    for r in dataframe_to_rows(st.session_state.df, index=False, header=True):
        sheet_3.append(r)
    img3 = Image('Зависимость ηол от U_cф.png')
    img4 = Image('new h - s диаграмма.png')
    img5 = Image('Треугольник скоростей.png')
    sheet_3.add_image(img3, 'F5')
    sheet_3.add_image(img4, 'Q1')
    sheet_3.add_image(img5, 'Q26')

    sheet_4 = newbook.create_sheet("Задание 3")
    for r in dataframe_to_rows(st.session_state.table, index=False, header=True):
        sheet_4.append(r)
    img6 = Image('Рисунок 1 Распределение средних диаметров по проточной части.png')
    img7 = Image('Рисунок 2 Распределение высот лопаток по проточной части.png')
    img8 = Image('Рисунок 3 Распределение обратной веерности по проточной части.png')
    img9 = Image('Рисунок 4 Распределение степени реактивности по проточной части.png')
    img10 = Image('Рисунок 5 Распределение U_Cф по проточной части.png')
    img11 = Image('Рисунок 6 Распределение теплоперепадов по проточной части.png')
    img12 = Image('Рисунок 7 Распределение теплоперепадов с учетом невязки по проточной части.png')

    sheet_4.add_image(img6, 'A25')
    sheet_4.add_image(img7, 'P25')
    sheet_4.add_image(img8, 'A50')
    sheet_4.add_image(img9, 'P50')
    sheet_4.add_image(img10, 'A75')
    sheet_4.add_image(img11, 'P75')
    sheet_4.add_image(img12, 'A100')

    newbook.save("Результаты " + sheet[st.session_state.index_row][0].value + ".xlsx")

def clear():
    os.remove("Зависимость КПД от давления пром. перегрева.png")
    os.remove("h - s диаграмма.png")
    os.remove("Треугольник скоростей.png")
    os.remove("new h - s диаграмма.png")
    os.remove("Зависимость ηол от U_cф.png")
    os.remove("Рисунок 1 Распределение средних диаметров по проточной части.png")
    os.remove("Рисунок 2 Распределение высот лопаток по проточной части.png")
    os.remove("Рисунок 3 Распределение обратной веерности по проточной части.png")
    os.remove("Рисунок 4 Распределение степени реактивности по проточной части.png")
    os.remove("Рисунок 5 Распределение U_Cф по проточной части.png")
    os.remove("Рисунок 6 Распределение теплоперепадов по проточной части.png")
    os.remove("Рисунок 7 Распределение теплоперепадов с учетом невязки по проточной части.png")
    time.sleep(1)
    os.remove("Результаты "+ sheet[st.session_state.index_row][0].value + ".xlsx")






st.write("Выполнено: Мурашов.В ФПэ-01-19")
st.write("Github: " + "https://github.com/tederix/PGT")


with st.sidebar:

    uploaded_file = st.file_uploader(label="Загрузите файл с заданиями", type='xlsx')
    if uploaded_file is not None:
        chek = False
        indexch=0
        book = openpyxl.open(uploaded_file, read_only=True)
        sheet = book.active
        FIO = []
        for row in range(2, sheet.max_row):
            A = sheet[row][0].value
            if A != None:
                FIO.append(A)
    else:
        st.write("Открыт базовый файл")
        chek = True
        indexch=8
        book = openpyxl.open("2022.xlsx", read_only=True)
        sheet = book.active
        FIO = []
        for row in range(2, sheet.max_row):
            A = sheet[row][0].value
            if A != None:
                FIO.append(A)
    st.session_state.sheet = sheet
    st.write("#")
    fio = st.selectbox(
        "Выберите вариант",
        (FIO), index=indexch)

    for row in range(2, sheet.max_row):
        if fio == sheet[row][0].value:
            index_row = row
            st.session_state.index_row = index_row
            break
    if chek:
        if(index_row ==4 or index_row ==10 or index_row ==18):
            st.write()
        else:
            st.subheader("Выберите варианты:\n Быковский, Мурашов или Буйницкий")

    page = st.selectbox(
        "Выберите задание",
        ("Задание 1", "Задание 2", "Задание 3", "Скачать результаты"))


if page == "Задание 1":

    st.write("""# """)
    st.write("# Задание 1")
    st.write("""Построить процесс расширения пара в турбине. Определение расходов пара на входе в турбину (G0) и в конденсатор (Gк). Получить зависимость КПД ПТУ от параметра заданного в таблице.""")
    st.write("""# """)

    st.write(" *Исходные данные:* ")

    Ne = st.number_input('Введите мощность Nэ, МВт', value=sheet[index_row][8].value) * 10 ** 6
    st.session_state.Ne = Ne

    p0 = st.number_input('Введите давление P0, МПа', value=sheet[index_row][2].value) * 10 ** 6
    st.session_state.p0 = p0/(10**6)

    t0 = st.number_input('Введите температуру T0, °C', value=sheet[index_row][3].value)
    st.session_state.t0 = t0
    T0 = t0 + 273.15

    tpp = st.number_input('Введите температуру Tпп, °C', value=sheet[index_row][5].value)
    Tpp = tpp + 273.15

    pk = st.number_input('Введите давление Pk, кПа', value=sheet[index_row][6].value) * 10 ** 3

    tpv = st.number_input('Введите температуру Tпв, °C', value=sheet[index_row][7].value)
    Tpv = tpv + 273.15


    age_min, age_max = sheet[index_row][4].value.split("-")
    age = st.slider('Укажите максимальную границу Pпп', min_value=float(age_min), max_value=float(age_max), step=0.1)
    age = age + 0.01
    P_pp = list(np.arange(float(age_min), age, 0.1))
    ppp = [p * 1e6 for p in P_pp]
    p_pp_min = float(ppp[0])
    p_pp_max = float(ppp[-1])

    delta_p_0 = 0.05 * p0
    delta_p_pp = 0.08 * p_pp_max
    delta_p = 0.03 * p_pp_max

    z = sheet[index_row][9].value
    st.session_state.z = z

    st.write("""# """)
    st.write(" *Дано:* ")
    st.write(""" P0 = """ + str(p0 * 10 ** (-6)) + """ МПа""")
    st.write(""" t0 = """ + str(t0) + """ C""")
    st.write(
        """ Pпп = """ + str(p_pp_min * 10 ** (-6)) + " - " + str('{:.2}'.format(p_pp_max * 10 ** (-6))) + """ МПа""")
    st.write(""" tпп = """ + str(tpp) + """ C """)
    st.write(""" Pк = """ + str(pk * 10 ** (-3)) + """ кПа """)
    st.write(""" tпв = """ + str(tpv) + """ C """)
    st.write(""" Nэ = """ + str(Ne * 10 ** (-6)) + """ МВт """)
    st.write(""" Z = """ + str(z) + """ шт """)


    st.write("""# """)
    st.write(" *Решение:* ")


    def Calculate_eta_G0_Gk(N_e, p_0, T_0, p_pp, T_pp, p_k, T_pv):

        point_0 = IAPWS97(P=p_0*10**(-6), T=T_0)
        s_0 = point_0.s
        h_0 = point_0.h
        v_0 = point_0.v
        p_0_ = p_0-0.05*p_0

        point_p_0_ = IAPWS97(P=p_0_*10**(-6), h=h_0)
        t_0_ = point_p_0_.T-273.15
        s_0_ = point_p_0_.s
        v_0_ = point_p_0_.v


        p_1t = p_pp+0.1*p_pp
        point_1t = IAPWS97(P=p_1t*10**(-6), s=s_0)
        t_1t = point_1t.T-273.15
        h_1t = point_1t.h
        v_1t = point_1t.v


        point_pp = IAPWS97 (P=p_pp*10**(-6), T=T_pp)
        h_pp = point_pp.h
        s_pp = point_pp.s
        v_pp = point_pp.v

        H_0 = h_0-h_1t
        eta_oi = 0.85
        H_i_cvd = H_0*eta_oi

        h_1 = h_0 - H_i_cvd
        point_1 = IAPWS97(P = p_1t*10**(-6),h = h_1)
        s_1 = point_1.s
        T_1 = point_1.T
        v_1 = point_1.v
        p_1 = point_1.P
        st.session_state.p_1 = f"{p_1:.2f}"

        p_pp_ = p_pp - 0.03*p_pp
        point_pp_ = IAPWS97(P=p_pp_*10**(-6),h = h_pp)
        s_pp_ = point_pp_.s
        v_pp_ = point_pp_.v
        point_kt = IAPWS97(P = p_k*10**(-6),s = s_pp)
        T_kt = point_kt.T
        h_kt = point_kt.h
        v_kt = point_kt.v
        s_kt = s_pp
        H_0_csdcnd = h_pp-h_kt
        eta_oi = 0.85
        H_i_csdcnd = H_0_csdcnd*eta_oi
        h_k = h_pp - H_i_csdcnd
        point_k = IAPWS97(P = p_k*10**(-6), h = h_k)
        T_k = point_k.T
        s_k = point_k.s
        v_k = point_k.v
        point_k_v = IAPWS97(P = p_k*10**(-6),x=0)
        h_k_v = point_k_v.h
        s_k_v = point_k_v.s
        eta_oiI = (h_1-h_0)/(h_1t-h_0)
        p_pv = 1.4*p_0
        point_pv = IAPWS97(P = p_pv*10**(-6),T=T_pv)
        h_pv = point_pv.h
        s_pv = point_pv.s
        ksi_pp_oo = 1-(1-(T_k*(s_pp-s_k_v))/((h_0-h_1t)+(h_pp-h_k_v)))/(1-(T_k*(s_pp-s_pv))/((h_0-h_1t)+(h_pp-h_pv)))
        #T_0_= IAPWS97(P = p_pv*10**(-6),x = 0).T
        T_0_ = 374.2+273.15
        T_ = (point_pv.T - point_k.T) / (T_0_ - point_k.T)
        if T_ <= 0.636364:
            ksi1 = -1.53*T_**2+2.1894*T_+0.0048
        elif 0.636364<T_<=0.736364:
            ksi1 = -1.3855*T_**2+2.0774*T_+0.0321
        elif 0.736364<T_<=0.863636:
            ksi1 = -2.6535*T_**2+4.2556*T_-0.8569


        if T_ <= 0.631818:
            ksi2 = -1.7131*T_**2+2.3617*T_-0.0142
        elif 0.631818<T_<=0.718182:
            ksi2 = -2.5821*T_**2+3.689*T_-0.4825
        elif 0.718182<T_<=0.827273:
            ksi2 = -1.9864*T_**2+3.138*T_-0.3626
        elif 0.827273<T_<=0.936364:
            ksi2 = -2.0619*T_**2+3.3818*T_-0.4814

        ksi = (ksi1+ksi2)/2
        ksi_r_pp = ksi*ksi_pp_oo
        eta_ir = (H_i_cvd+H_i_csdcnd)/(H_i_cvd+(h_pp-h_k_v))*1/(1-ksi_r_pp)
        H_i = eta_ir*((h_0-h_pv)+(h_pp-h_1))
        eta_m = 0.994
        eta_eg = 0.99
        G_0 = N_e/(H_i*eta_m*eta_eg*(10**3))
        G_k = N_e/((h_k-h_k_v)*eta_m*eta_eg*(10**3))*(1/eta_ir-1)

        return eta_ir, G_0, G_k

    eta, G0, Gk =[], [], []
    for p in ppp:
        eta_ = Calculate_eta_G0_Gk(N_e = Ne, p_0 = p0, T_0 = T0, p_pp = p, T_pp = Tpp, p_k = pk, T_pv = Tpv)
        eta.append(eta_[0])
        G0.append(eta_[1])
        Gk.append(eta_[2])

    max: float = eta[0]
    pos = 0
    for i in range(len(eta)):
        if eta[i] > max: max = eta[i]; pos = i


    ppp_f = [float(x) * 10**(-6) for x in ppp]
    st.session_state.ppp_fmax=ppp_f[pos]
    eta_f = [float(x) * 100 for x in eta]

    st.write(""" Максимальный КПД = """ + str('{:.4}'.format(float(eta_f[pos]))) + """ %""")
    st.write(""" Расход пара на входе в турбину (G0) при макс. КПД = """ + str('{:.5}'.format(float(G0[pos]))) + """ кг/с""")
    st.write(""" Расход пара на входе в конденсатор (Gк) при макс. КПД = """ + str('{:.5}'.format(float(Gk[pos]))) + """ кг/с""")
    st.session_state.eta_fmax =float(str('{:.4}'.format(float(eta_f[pos]))))
    st.session_state.G0max = float(str('{:.5}'.format(float(G0[pos]))))
    st.session_state.Gkmax =float(str('{:.5}'.format(float(Gk[pos]))))
    st.session_state.G0 = f"{G0[pos]:.4f}"

    st.write("""# """)
    st.write(" Табл. Зависимость КПД от Pпп  ")


    ppp_eta=pd.DataFrame({"ppp, МПа": (ppp_f),
                       "eta, %": (eta_f),
                       "G_0, кг/с": (G0),
                       "G_k, кг/с": (Gk)
                       })
    st.dataframe(ppp_eta)
    st.session_state.ppp_eta=ppp_eta

    st.write("""# """)

    ppp__eta = plt.figure()

    plt.plot(ppp_f, eta_f)
    plt.plot(ppp_f, eta_f, 'ro')
    plt.title("Зависимость КПД от давления пром. перегрева")
    plt.xlabel("P_пп, MПа")
    plt.ylabel("КПД, %")
    plt.grid()
    plt.savefig('Зависимость КПД от давления пром. перегрева.png')
    st.pyplot(ppp__eta)







    st.title(""" """)

    p_pp_max = ppp[pos]

    fighs = plt.figure()

    point_0 = IAPWS97(P=p0*1e-6, T=T0)
    p_0_d = p0 - delta_p_0
    st.session_state.p_0_d = f"{p_0_d/(10**6):.2f}"
    point_0_d = IAPWS97(P=p_0_d*1e-6, h=point_0.h)
    st.session_state.h_0_d = f"{point_0_d.h:.2f}"
    p_1t = p_pp_max + delta_p_pp
    point_1t = IAPWS97(P=p_1t*10**(-6), s=point_0.s)
    H_01 = point_0.h - point_1t.h
    kpd_oi = 0.85
    H_i_cvd = H_01 * kpd_oi
    h_1 = point_0.h - H_i_cvd
    point_1 = IAPWS97(P=p_1t*1e-6, h=h_1)
    point_pp = IAPWS97(P=p_pp_max*1e-6, T=Tpp)
    p_pp_d = p_pp_max - delta_p_pp
    point_pp_d = IAPWS97(P=p_pp_d*1e-6, h=point_pp.h)
    point_kt = IAPWS97(P=pk*1e-6, s=point_pp.s)
    H_02 = point_pp.h - point_kt.h
    kpd_oi = 0.85
    H_i_csd_cnd = H_02 * kpd_oi
    h_k = point_pp.h - H_i_csd_cnd
    point_k = IAPWS97(P=pk*1e-6, h=h_k)

    s_0 = [point_0.s-0.05,point_0.s,point_0.s+0.05]
    h_0 = [IAPWS97(P = p0*1e-6,s = s_).h for s_ in s_0]
    s_1 = [point_0.s-0.05,point_0.s,point_0.s+0.18]
    h_1 = [IAPWS97(P=p_1t*1e-6, s = s_).h for s_ in s_1]
    s_0_d = [point_0_d.s-0.05, point_0_d.s, point_0_d.s+0.05]
    h_0_d = h_0
    s_pp = [point_pp.s-0.05,point_pp.s,point_pp.s+0.05]
    h_pp = [IAPWS97(P=p_pp_max*1e-6, s=s_).h for s_ in s_pp]
    s_k = [point_pp.s-0.05,point_pp.s,point_pp.s+0.8]
    h_k = [IAPWS97(P=pk*1e-6, s=s_).h for s_ in s_k]
    s_pp_d = [point_pp_d.s-0.05,point_pp_d.s,point_pp_d.s+0.05]
    h_pp_d = h_pp

    plt.plot([point_0.s,point_0.s,point_0_d.s,point_1.s],[point_1t.h,point_0.h,point_0.h,point_1.h],'-or')
    plt.plot([point_pp.s,point_pp.s,point_pp_d.s,point_k.s],[point_kt.h,point_pp.h,point_pp.h,point_k.h],'-or')
    plt.plot(s_0,h_0)
    plt.plot(s_1,h_1)
    plt.plot(s_0_d,h_0_d)
    plt.plot(s_pp,h_pp)
    plt.plot(s_k,h_k)
    plt.plot(s_pp_d,h_pp_d)

    for x, y, ind in zip([point_pp.s, point_k.s], [point_pp.h, point_k.h], ['{пп}', '{к}']):
        plt.text(x-0.45, y+40, '$h_' + ind + ' = %.2f $'%y)
    for x, y, ind in zip([point_kt.s, point_pp_d.s], [point_kt.h, point_pp_d.h], ['{кт}', '{ппд}']):
        plt.text(x+0.03, y+40, '$h_' + ind + ' = %.2f $'%y)

    for x, y, ind in zip ([point_0.s, point_1.s], [point_0.h, point_1.h], ['{0}', '{1}']):
        plt.text(x-0.01, y+120, '$h_' + ind + ' = %.2f $'%y)

    for x, y, ind in zip([point_1t.s, point_0_d.s], [point_1t.h, point_0_d.h], ['{1т}', '{0д}']):
        plt.text(x+0.03, y-60, '$h_' + ind + ' = %.2f $'%y)


        plt.title("h - s диаграмма")
        plt.xlabel("s, кДж/(кг*С)")
        plt.ylabel("h, кДж/кг")
        plt.grid(True)

    plt.savefig('h - s диаграмма.png')
    st.pyplot(fighs)

if page == "Задание 2":

    st.write("# Задание 2")

    st.write("""Проведение расчета регулирующей ступени и определение зависимости ηол от U/cф.""")
    st.write("""# """)

    st.write(" *Исходные данные:* ")

    d = 1.1  # m
    n = sheet[index_row][11].value  # Гц
    st.session_state.n = n
    rho = 0.05
    l_1 = 0.0265  # м
    alpha_1 = 12  # град
    b_1 = 0.06  # м
    Delta = 0.003  # м
    b_2 = 0.03  # м
    kappa_vs = 0  # коэф исп вых скорости

    Ne = st.session_state.Ne
    p_0 = st.session_state.p0
    t_0 = st.session_state.t0
    T_0 = t_0 + 273.15
    G_0=float(st.session_state.G0)
    H_0 = st.number_input('Введите теплоперепад H0, кДж/кг', value=float(sheet[index_row][12].value))
    z = st.session_state.z

    st.write("""# """)
    st.write(" *Дано:* ")
    st.write(""" Nэ = """ + str(Ne * 10 ** (-6)) + """ МВт """)
    st.write(""" P0 = """ + str(p_0) + """ МПа""")
    st.write(""" t0 = """ + str(t_0) + """ C""")
    st.write(""" G0 = """ + str(G_0) + """ кг/с """)
    st.write(""" H0 = """ + str(H_0) + """ кДж/кг """)
    st.write(""" dp.c. = """ + str(0.9) + " - " + str(1.1) + """ м """)
    st.write(""" Z = """ + str(z) + """ шт """)
    st.write(""" n = """ + str(n) + """ Гц """)

    st.write("""# """)
    st.write(" *Решение:* ")


    def iso_bar(wsp_point, min_s=-0.1, max_s=0.11, step_s=0.011, color='r'):
        if not isinstance(wsp_point, list):
            iso_bar_0_s = np.arange(wsp_point.s + min_s, wsp_point.s + max_s, step_s).tolist()
            iso_bar_0_h = [WSP(P=wsp_point.P, s=i).h for i in iso_bar_0_s]
        else:
            iso_bar_0_s = np.arange(wsp_point[0].s + min_s, wsp_point[1].s + max_s, step_s).tolist()
            iso_bar_0_h = [WSP(P=wsp_point[1].P, s=i).h for i in iso_bar_0_s]
        plt.plot(iso_bar_0_s, iso_bar_0_h, color)

    def callculate_optimum(d, p_0, T_0, n, G_0, H_0, rho, l_1, alpha_1, b_1, Delta, b_2, kappa_vs):
        u = M.pi * d * n
        point_0 = WSP(P=p_0, T=T_0)
        H_0s = H_0 * (1 - rho)
        H_0r = H_0 * rho
        h_1t = point_0.h - H_0s
        point_1t = WSP(h=h_1t, s=point_0.s)
        c_1t = (2000 * H_0s) ** 0.5
        M_1t = c_1t / point_1t.w
        mu_1 = 0.982 - 0.005 * (b_1 / l_1)
        F_1 = G_0 * point_1t.v / mu_1 / c_1t
        el_1 = F_1 / M.pi / d / M.sin(M.radians(alpha_1))
        e_opt = 5 * el_1 ** 0.5
        if e_opt > 0.85:
            e_opt = 0.85
        l_1 = el_1 / e_opt

        fi_1 = 0.98 - 0.008 * (b_1 / l_1)
        c_1 = c_1t * fi_1
        alpha_1 = M.degrees(M.asin(mu_1 / fi_1 * M.sin(M.radians(alpha_1))))
        w_1 = (c_1 ** 2 + u ** 2 - 2 * c_1 * u * M.cos(M.radians(alpha_1))) ** 0.5
        betta_1 = M.degrees(M.atan(M.sin(M.radians(alpha_1)) / (M.cos(M.radians(alpha_1)) - u / c_1)))
        Delta_Hs = c_1t ** 2 / 2 * (1 - fi_1 ** 2)
        h_1 = h_1t + Delta_Hs * 1e-3
        point_1 = WSP(P=point_1t.P, h=h_1)
        h_2t = h_1 - H_0r
        point_2t = WSP(h=h_2t, s=point_1.s)
        w_2t = (2 * H_0r * 1e3 + w_1 ** 2) ** 0.5
        l_2 = l_1 + Delta
        mu_2 = 0.965 - 0.01 * (b_2 / l_2)
        M_2t = w_2t / point_2t.w
        F_2 = G_0 * point_2t.v / mu_2 / w_2t
        betta_2 = M.degrees(M.asin(F_2 / (e_opt * M.pi * d * l_2)))
        point_1w = WSP(h=point_1.h + w_1 ** 2 / 2 * 1e-3, s=point_1.s)

        psi = 0.96 - 0.014 * (b_2 / l_2)
        w_2 = psi * w_2t
        c_2 = (w_2 ** 2 + u ** 2 - 2 * u * w_2 * M.cos(M.radians(betta_2))) ** 0.5
        alpha_2 = M.degrees(M.atan(M.sin(M.radians(betta_2)) / (M.cos(M.radians(betta_2)) - u / w_2)))
        if alpha_2 < 0:
            alpha_2 = 180 + alpha_2
        Delta_Hr = w_2t ** 2 / 2 * (1 - psi ** 2)
        h_2 = h_2t + Delta_Hr * 1e-3
        point_2 = WSP(P=point_2t.P, h=h_2)
        Delta_Hvs = c_2 ** 2 / 2
        E_0 = H_0 - kappa_vs * Delta_Hvs
        etta_ol1 = (E_0 * 1e3 - Delta_Hs - Delta_Hr - (1 - kappa_vs) * Delta_Hvs) / (E_0 * 1e3)
        etta_ol2 = (u * (c_1 * M.cos(M.radians(alpha_1)) + c_2 * M.cos(M.radians(alpha_2)))) / (E_0 * 1e3)
        return etta_ol2, alpha_2

    d_min, d_max = sheet[index_row][10].value.replace(',', '.').split("-")
    d = [i * 1e-2 for i in list(range(int((float(d_min))*100), int((float(d_max))*100+1), 1))]
    alpha1 = []
    eta = []
    ucf = []


    fighs = plt.figure()
    for i in d:
        ucf_1 = M.pi * i * n / (2000 * H_0) ** 0.5
        ucf.append(ucf_1)

        eta_ol, alpha = callculate_optimum(i, p_0, T_0, n, G_0, H_0, rho, l_1, alpha_1, b_1, Delta, b_2, kappa_vs)
        alpha1.append(alpha)
        eta.append(eta_ol)
    plt.plot(ucf, eta)
    plt.ylabel('ηол')
    plt.xlabel('U/cф')
    plt.title("Зависимость ηол от U/cф")
    plt.grid(True)
    plt.savefig('Зависимость ηол от U_cф.png')
    st.pyplot(fighs)


    def frange(x, y, jump):
        while x < y:
            yield x
            x += jump

    st.write("""# """)
    st.write("Табл. Зависимость ηол от U/cф ")
    df = pd.DataFrame({
        "d, м": list(frange(float(d_min), float(d_max)+0.01, 0.01)),
        "ηол": (eta),
        "α": (alpha1),
        "U/cf": (ucf)})      #Таблица
    df
    st.session_state.df=df
    st.write("""# """)


    d = 1.1
    u = M.pi * d * n

    point_0 = WSP(P=p_0, T=T_0)
    H_0s = H_0 * (1 - rho)
    H_0r = H_0 * rho
    h_1t = point_0.h - H_0s
    point_1t = WSP(h=h_1t, s=point_0.s)
    c_1t = (2000 * H_0s) ** 0.5
    M_1t = c_1t / point_1t.w
    mu_1 = 0.982 - 0.005 * (b_1 / l_1)
    F_1 = G_0 * point_1t.v / mu_1 / c_1t
    el_1 = F_1 / M.pi / d / M.sin(M.radians(alpha_1))
    e_opt = 6 * el_1 ** 0.5
    if e_opt > 0.85:
        e_opt = 0.85
    l_1 = el_1 / e_opt

    #st.write(f'u = {u:.2f} м/с')
    #st.write(f'h_0 = {point_0.h:.2f} кДж/кг')
    #st.write(f's_0 = {point_0.s:.4f} кДж/(кг*К)')
    #st.write(f'h_1т = {h_1t:.2f} кДж/кг')
    #st.write(f'c_1т = {c_1t:.2f} м/с')
    # st.write(f'M_1т = {M_1t:.2f}')
    # st.write(f'F_1 = {F_1:.4f} м^2')
    # st.write(f'el_1 = {el_1:.4f} м')
    #st.write(f'l_1 = {l_1:.4f} м')

    fignozzle = plt.figure()
    def plot_hs_nozzle_t(x_lim, y_lim):
        plt.plot([point_0.s, point_1t.s], [point_0.h, point_1t.h], 'ro-')
        iso_bar(point_0, -0.02, 0.02, 0.001, 'c')
        iso_bar(point_1t, -0.02, 0.02, 0.001, 'y')
        plt.xlim(x_lim)
        plt.ylim(y_lim)
    plot_hs_nozzle_t([6.1, 6.5], [3300, 3600])
    plt.ylabel('h кДж/кг')
    plt.xlabel('s кДж/кг*К')
    #st.pyplot(fignozzle)


    #st.write(f'l_1 = {l_1:.4f} м')

    if alpha_1 <= 10:
        NozzleBlade = 'C-90-09A'
        t1_ = 0.78
        b1_mod = 6.06
        f1_mod = 3.45
        W1_mod = 0.471
        alpha_inst1 = alpha_1 - 12.5 * (t1_ - 0.75) + 20.2
    elif 10 < alpha_1 <= 13:
        NozzleBlade = 'C-90-12A'
        t1_ = 0.78
        b1_mod = 5.25
        f1_mod = 4.09
        W1_mod = 0.575
        alpha_inst1 = alpha_1 - 10 * (t1_ - 0.75) + 21.2
    elif 13 < alpha_1 <= 16:
        NozzleBlade = 'C-90-15A'
        t1_ = 0.78
        b1_mod = 5.15
        f1_mod = 3.3
        W1_mod = 0.45
        alpha_inst1 = alpha_1 - 16 * (t1_ - 0.75) + 23.1
    else:
        NozzleBlade = 'C-90-18A'
        t1_ = 0.75
        b1_mod = 4.71
        f1_mod = 2.72
        W1_mod = 0.333
        alpha_inst1 = alpha_1 - 17.7 * (t1_ - 0.75) + 24.2

    #st.write('Тип профиля:', NozzleBlade)
    #st.write(f'Оптимальный относительный шаг t1_ = {t1_}')
    z1 = (M.pi * d) / (b_1 * t1_)
    z1 = int(z1)
    #if z1 % 2 == 0:
        #st.write(f'z1 = {z1}')
    #else:
        #z1 = z1 + 1

        #st.write(f'z1 = {z1}')

    t1_ = (M.pi * d) / (b_1 * z1)
    Ksi_1_ = (0.021042 * b_1 / l_1 + 0.023345) * 100
    k_11 = 7.18977510 * M_1t ** 5 - 26.94497258 * M_1t ** 4 + 39.35681781 * M_1t ** 3 - 26.09044664 * M_1t ** 2 + 6.75424811 * M_1t + 0.69896998
    k_12 = 0.00014166 * 90 ** 2 - 0.03022881 * 90 + 2.61549380
    k_13 = 13.25474043 * t1_ ** 2 - 20.75439502 * t1_ + 9.12762245
    Ksi_1 = Ksi_1_ * k_11 * k_12 * k_13

    fi_1 = M.sqrt(1 - Ksi_1 / 100)

    #st.write(f'mu_1 = {mu_1}')
    #st.write(f'fi_1 = {fi_1}')

    alpha_1 = 12
    c_1 = c_1t * fi_1

    alpha_1 = M.degrees(M.asin(mu_1 / fi_1 * M.sin(M.radians(alpha_1))))

    w_1 = (c_1 ** 2 + u ** 2 - 2 * c_1 * u * M.cos(M.radians(alpha_1))) ** 0.5

    #st.write(f'c_1 = {c_1:.2f} м/с')
    #st.write(f'alpha_1 = {alpha_1:.2f} град.')
    #st.write(f'w_1 = {w_1}')
    c_1u = c_1 * M.cos(M.radians(alpha_1))
    c_1a = c_1 * M.sin(M.radians(alpha_1))
    w_1u = c_1u - u

    #st.write(c_1u, w_1u)
    w_1_tr = [0, 0, -w_1u, -c_1a]
    c_1_tr = [0, 0, -c_1u, -c_1a]
    u_1_tr = [-w_1u, -c_1a, -u, 0]


    fig2 = plt.figure()
    ax = plt.axes()
    ax.arrow(*c_1_tr, head_width=5, length_includes_head=True, head_length=20, fc='r', ec='r')
    ax.arrow(*w_1_tr, head_width=5, length_includes_head=True, head_length=20, fc='b', ec='b')
    ax.arrow(*u_1_tr, head_width=5, length_includes_head=True, head_length=20, fc='g', ec='g')
    plt.text(-2 * c_1u / 3, -3 * c_1a / 4, '$c_1$', fontsize=20)
    plt.text(-2 * w_1u / 3, -3 * c_1a / 4, '$w_1$', fontsize=20)
    #st.pyplot(fig2)


    betta_1 = M.degrees(M.atan(M.sin(M.radians(alpha_1)) / (M.cos(M.radians(alpha_1)) - u / c_1)))
    Delta_Hs = c_1t ** 2 / 2 * (1 - fi_1 ** 2)
    h_1 = h_1t + Delta_Hs * 1e-3
    point_1 = WSP(P=point_1t.P, h=h_1)
    h_2t = h_1 - H_0r
    point_2t = WSP(h=h_2t, s=point_1.s)
    w_2t = (2 * H_0r * 1e3 + w_1 ** 2) ** 0.5
    l_2 = l_1 + Delta
    mu_2 = 0.965 - 0.01 * (b_2 / l_2)
    M_2t = w_2t / point_2t.w
    F_2 = G_0 * point_2t.v / mu_2 / w_2t
    betta_2 = M.degrees(M.asin(F_2 / (e_opt * M.pi * d * l_2)))
    point_1w = WSP(h=point_1.h + w_1 ** 2 / 2 * 1e-3, s=point_1.s)
    #st.write(f'betta_1 = {betta_1:.2f}')
    #st.write(f'Delta_Hs = {Delta_Hs:.2f} Дж/кг')
    #st.write(f'h_1 = {h_1:.2f} кДж/кг')
    #st.write(f'h_2t = {h_2t:.2f} кДж/кг')
    #st.write(f'w_2t = {w_2t:.2f} м/с')
    #st.write(f'mu_2 = {mu_2:.2f}')
    #st.write(f'M_2t = {M_2t:.2f}')
    #st.write(f'F_2 = {F_2:.2f}')
    #st.write(f'betta_2 = {betta_2:.2f}')



    if betta_2 <= 15:
        RotorBlade = 'P-23-14A'
        t2_ = 0.63
        b2_mod = 2.59
        f2_mod = 2.44
        W2_mod = 0.39
        beta_inst2 = betta_2 - 12.5 * (t2_ - 0.75) + 20.2
    elif 15 < betta_2 <= 19:
        RotorBlade = 'P-26-17A'
        t2_ = 0.65
        b2_mod = 2.57
        f2_mod = 2.07
        W2_mod = 0.225
        beta_inst2 = betta_2 - 19.3 * (t2_ - 0.6) + 60
    elif 19 < betta_2 <= 23:
        RotorBlade = 'P-30-21A'
        t2_ = 0.63
        b2_mod = 2.56
        f2_mod = 1.85
        W2_mod = 0.234
        beta_inst2 = betta_2 - 12.8 * (t2_ - 0.65) + 58
    elif 23 < betta_2 <= 27:
        RotorBlade = 'P-35-25A'
        t2_ = 0.6
        b2_mod = 2.54
        f2_mod = 1.62
        W2_mod = 0.168
        beta_inst2 = betta_2 - 16.6 * (t2_ - 0.65) + 54.3
    elif 27 < betta_2 <= 31:
        RotorBlade = 'P-46-29A'
        t2_ = 0.51
        b2_mod = 2.56
        f2_mod = 1.22
        W2_mod = 0.112
        beta_inst2 = betta_2 - 50.5 * (t2_ - 0.6) + 47.1
    else:
        RotorBlade = 'P-50-33A'
        t2_ = 0.49
        b2_mod = 2.56
        f2_mod = 1.02
        W2_mod = 0.079
        beta_inst2 = betta_2 - 20.8 * (t2_ - 0.6) + 43.7

    z2 = int((M.pi * d) / (b_2 * t2_))

    t2_ = (M.pi * d) / (b_2 * z2)
    Ksi_2_ = 4.364 * b_2 / l_2 + 4.22
    k_21 = -13.79438991 * M_2t ** 4 + 36.69102267 * M_2t ** 3 - 32.78234341 * M_2t ** 2 + 10.61998662 * M_2t + 0.28528786
    k_22 = 0.00331504 * betta_1 ** 2 - 0.21323910 * betta_1 + 4.43127194
    k_23 = 60.72813684 * t2_ ** 2 - 76.38053189 * t2_ + 24.97876023
    Ksi_2 = Ksi_2_ * k_21 * k_22 * k_23

    psi = M.sqrt(1 - Ksi_2 / 100)

    psi = 0.93

    w_2 = psi * w_2t

    c_2 = (w_2 ** 2 + u ** 2 - 2 * u * w_2 * M.cos(M.radians(betta_2))) ** 0.5

    alpha_2 = M.degrees(M.atan(M.sin(M.radians(betta_2)) / (M.cos(M.radians(betta_2)) - u / w_2)))

    Delta_Hr = w_2t ** 2 / 2 * (1 - psi ** 2)

    h_2 = h_2t + Delta_Hr * 1e-3
    point_2 = WSP(P=point_2t.P, h=h_2)
    Delta_Hvs = c_2 ** 2 / 2

    E_0 = H_0 - kappa_vs * Delta_Hvs
    etta_ol1 = (E_0 * 1e3 - Delta_Hs - Delta_Hr - (1 - kappa_vs) * Delta_Hvs) / (E_0 * 1e3)

    etta_ol2 = (u * (c_1 * M.cos(M.radians(alpha_1)) + c_2 * M.cos(M.radians(alpha_2)))) / (E_0 * 1e3)


    h_vs = h_2 + Delta_Hvs * 1e-3
    point_vs = IAPWS97(P=point_2t.P, h=h_vs)


    fig3 = plt.figure()
    def plot_hs_stage_t(x_lim, y_lim):
        plot_hs_nozzle_t(x_lim, y_lim)
        plt.plot([point_0.s, point_1.s], [point_0.h, point_1.h], 'bo-')
        plt.plot([point_1.s, point_2t.s], [point_1.h, point_2t.h], 'ro-')
        #plt.plot([point_1.s, point_1.s], [point_1w.h, point_1.h], 'ro-')
        iso_bar(point_2t, -0.02, 0.02, 0.001, 'y')
        plt.plot([point_2.s, point_vs.s], [point_2.h, point_vs.h], 'ro-')
        plt.plot([point_1.s, point_2.s], [point_1.h, point_2.h], 'bo-')
        #iso_bar(point_2t, -0.02, 0.02, 0.001, 'y')
        #iso_bar(point_1w, -0.005, 0.005, 0.001, 'c')
    plt.ylabel('h кДж/кг')
    plt.xlabel('s кДж/кг*К')
    plt.title("h - s диаграмма")
    plot_hs_stage_t([point_0.s - 0.005, point_vs.s + 0.005], [point_2t.h - 10, point_0.h + 10])
    plt.grid(True)
    plt.savefig('new h - s диаграмма.png')
    st.pyplot(fig3)









    c_1u = c_1 * M.cos(M.radians(alpha_1))
    c_1a = c_1 * M.sin(M.radians(alpha_1))
    w_1u = c_1u - u
    w_2a = w_2 * M.sin(M.radians(betta_2))
    w_2u = w_2 * M.cos(M.radians(betta_2))
    c_2u = w_2u + u
    print(c_1u, w_1u)
    w_1_tr = [0, 0, -w_1u, -c_1a]
    c_1_tr = [0, 0, -c_1u, -c_1a]
    u_1_tr = [-w_1u, -c_1a, -u, 0]

    #st.write('Тип профиля:', RotorBlade)
    #st.write(f'Оптимальный относительный шаг t2_ = {t2_}')
    #st.write(f'z2 = {z2}')
    #st.write(f'psi = {psi:.2f}')
    #st.write(f'w_2 = {w_2:.2f} м/с')
    #st.write(f'c_2 = {c_2:.2f} м/с')
    #st.write(f'alpha_2 = {alpha_2:.2f}')
    #st.write(f'Delta_Hr = {Delta_Hr:.2f} Дж/кг')
    #st.write(f'Delta_Hvs = {Delta_Hvs:.2f} Дж/кг')
    #st.write(f'1. etta_ol = {etta_ol1}')
    #st.write(f'2. etta_ol = {etta_ol2}')

    w_2_tr = [0, 0, w_2u, -w_2a]
    c_2_tr = [0, 0, c_2u, -w_2a]
    u_2_tr = [c_2u, -w_2a, -u, 0]


    fig4 = plt.figure()
    ax = plt.axes()
    ax.arrow(*c_1_tr, head_width=5, length_includes_head=True, head_length=20, fc='r', ec='r')
    ax.arrow(*w_1_tr, head_width=5, length_includes_head=True, head_length=20, fc='b', ec='b')
    ax.arrow(*u_1_tr, head_width=5, length_includes_head=True, head_length=20, fc='g', ec='g')
    ax.arrow(*c_2_tr, head_width=5, length_includes_head=True, head_length=20, fc='r', ec='r')
    ax.arrow(*w_2_tr, head_width=5, length_includes_head=True, head_length=20, fc='b', ec='b')
    ax.arrow(*u_2_tr, head_width=5, length_includes_head=True, head_length=20, fc='g', ec='g')
    plt.text(-2 * c_1u / 3, -3 * c_1a / 4, '$c_1$', fontsize=20)
    plt.text(-2 * w_1u / 3, -3 * c_1a / 4, '$w_1$', fontsize=20)
    plt.text(2.5 * c_2u / 3, -3 * w_2a / 4, '$c_2$', fontsize=20)
    plt.text(2.5 * w_2u / 3, -3 * w_2a / 4, '$w_2$', fontsize=20)
    plt.title("Треугольник скоростей")
    plt.savefig('Треугольник скоростей.png')
    st.pyplot(fig4)


    delta_a = 0.0025
    z_per_up = 2
    mu_a = 0.5
    mu_r = 0.75
    d_per = d + l_1
    delta_r = d_per * 0.001
    delta_ekv = 1 / M.sqrt(1 / (mu_a * delta_a) ** 2 + z_per_up / (mu_r * delta_r) ** 2)
    xi_u_b = M.pi * d_per * delta_ekv * etta_ol1 / F_1 * M.sqrt(rho + 1.8 * l_2 / d)
    Delta_Hub = xi_u_b * E_0
    k_tr = 0.0007
    Kappa_VS = 0
    u = M.pi * d * n
    c_f = M.sqrt(2000 * H_0)
    ucf = u / c_f
    xi_tr = k_tr * d ** 2 / F_1 * ucf ** 3
    Delta_Htr = xi_tr * E_0
    k_v = 0.065
    m = 1
    xi_v = k_v / M.sin(M.radians(alpha_1)) * (1 - e_opt) / e_opt * ucf ** 3 * m
    i_p = 4
    B_2 = b_2 * M.sin(M.radians(beta_inst2))
    xi_segm = 0.25 * B_2 * l_2 / F_1 * ucf * etta_ol1 * i_p
    xi_parc = xi_v + xi_segm
    Delta_H_parc = E_0 * xi_parc
    H_i = E_0 - Delta_Hr * 1e-3 - Delta_Hs * 1e-3 - (1 - Kappa_VS) * Delta_Hvs * 1e-3 - Delta_Hub - Delta_Htr - Delta_H_parc
    eta_oi = H_i / E_0
    N_i = G_0 * H_i

    #st.write("""Эквивалентный зазор в уплотнении по бандажу (периферийном) delta_ekv = %.3f мм""" % (delta_ekv * 1000))
    #st.write("""Относительные потери от утечек через бандажные уплотнения xi_u_b = %.3f""" % xi_u_b)
    #st.write("""Абсолютные потери от утечек через периферийное уплотнение ступени  Delta_Hub = %.3f кДж/кг""" % Delta_Hub)
    #st.write("""Определяем u/c_ф для ступени  U/c_ф = %.3f""" % ucf)
    #st.write("""Относительные потери от трения диска  xi_tr = %.5f""" % xi_tr)
    #st.write("""Абсолютные потери от трения диска  Delta_Htr = %.3f кДж/кг""" % Delta_Htr)
    #st.write("""Относительные вентиляционные потери""", xi_v)
    #st.write("""Относительные сегментные потери""", xi_segm)
    #st.write("""Использованный теплоперепад ступени  H_i = %.3f кДж/кг""" % H_i)

    st.write("""# """)
    st.write("""Внутренний относительный КПД ступени   eta_oi = %.3f """ % eta_oi)
    st.write("""Внутреняя мощность ступени  N_i = %.2f кВт""" % N_i)
    st.session_state.eta_oi=f"{eta_oi:.4f}"
    st.session_state.N_i = f"{N_i:.2f}"

if page == "Задание 3":


    st.write("# Задание 3")

    st.write("""Определение числа ступеней и распределение параметров по ним.""")
    st.write("""# """)

    st.write(" *Исходные данные:* ")


    P0 = float(st.session_state.p_0_d)
    h0 = float(st.session_state.h_0_d)
    Pz = float(st.session_state.p_1)
    G0 = float(st.session_state.G0)
    drs = 1.1
    etaoi = float(st.session_state.eta_oi)




    deltaD = st.slider('Разница между диаметром первой нерегулируемой ступени и регулирующей ступени: Δ𝑑', min_value=0.2, max_value=0.26, step=0.01) #m
    n = st.session_state.n  # Гц
    rho_s = st.slider('Степень реактивности первой нерегулируемой ступени в корне: 𝜌к', min_value=0.03, max_value=0.07, step=0.01)
    alfa = st.slider('Эффективный угол выхода потока из сопловой решетки: 𝛼1эф', min_value=10, max_value=16, step=1)  # град
    fi = st.slider('Коэффициент скорости сопловой решетки: 𝜙', min_value=0.93, max_value=0.96, step=0.01)
    mu1 = st.slider('Коэффициент расхода сопловой решетки первой нерегулируемой ступени: 𝜇', min_value=0.95, max_value=0.97, step=0.01)
    delta = 0.003
    tetta = 20
    Z=st.session_state.z

    st.write("""# """)
    st.write(" *Дано:* ")
    st.write(""" P0 = """ + str(P0) + """ МПа""")
    st.write(""" h0 = """ + str(h0) + """ кДж/кг""")
    st.write(""" Pz = """ + str(Pz) + """ МПа """)
    st.write(""" G0 = """ + str(G0) + """ кг/с """)
    st.write(""" dрс = """ + str(drs) + """ м """)
    st.write(""" eta_oi = """ + str(etaoi) + """ """)
    st.write(""" n = """ + str(n) + """ Гц """)

    st.write(""" Δ𝑑 = """ + str(deltaD) + """ м""")
    st.write(""" 𝜌к = """ + str(rho_s) + """""")
    st.write(""" 𝛼1эф = """ + str(alfa) + """ град""")
    st.write(""" 𝜙 = """ + str(fi) + """""")
    st.write(""" 𝜇 = """ + str(drs) + """ м """)





    st.write("""# """)
    st.write(" *Решение:* ")


    D1 = drs - deltaD
    sat_steam = IAPWS97(P=P0, h=h0)
    s_0 = sat_steam.s
    t_0 = sat_steam.T

    error = 2
    i = 1
    while error > 0.5:
        rho = rho_s + 1.8 / (tetta + 1.8)
        X = (fi * M.cos(M.radians(alfa))) / (2 * M.sqrt(1 - rho))
        H01 = 12.3 * (D1 / X) ** 2 * (n / 50) ** 2
        h2t = h0 - H01
        steam2t = IAPWS97(h=h2t, s=s_0)
        v2t = steam2t.v
        l11 = G0 * v2t * X / (M.pi ** 2 * D1 ** 2 * n * M.sqrt(1 - rho) * M.sin(M.radians(alfa)) * mu1)
        tetta_old = tetta
        tetta = D1 / l11
        #print(i, tetta_old, tetta)
        error = abs(tetta - tetta_old) / tetta_old * 100
        #print(error)
        i += 1

    l21 = l11 + delta
    d_s = D1 - l21
    steam_tz = IAPWS97(P=Pz, s=s_0)
    h_zt = steam_tz.h
    H0 = h0 - h_zt
    Hi = H0 * etaoi
    h_z = h0 - Hi
    steam_z = IAPWS97(P=Pz, h=h_z)
    v_2z = steam_z.v
    x = Symbol('x')
    с = solve(x ** 2 + x * d_s - (l21 * (d_s + l21) * v_2z / v2t))
    for j in с:
        if j > 0:
            l2z = j
    d2z = d_s + l2z
    tetta1 = (l21 + d_s) / l21
    tettaz = (l2z + d_s) / l2z
    rho1 = rho_s + 1.8 / (1.8 + tetta1)
    rhoz = rho_s + 1.8 / (1.8 + tettaz)
    X1 = (fi * cos(M.radians(alfa))) / (2 * sqrt(1 - rho1))
    Xz = (fi * cos(M.radians(alfa))) / (2 * sqrt(1 - rhoz))

    DeltaZ = 1
    ite = 0
    while DeltaZ > 0:
        matr = []
        Num = 0
        SumH = 0
        for _ in range(int(Z)):
            li = (l21 - l2z) / (1 - Z) * Num + l21
            di = (D1 - d2z) / (1 - Z) * Num + D1
            tetta_i = di / li
            rho_i = rho_s + 1.8 / (1.8 + tetta_i)
            X_i = (fi * M.cos(M.radians(alfa))) / (2 * M.sqrt(1 - rho_i))
            if Num < 1:
                H_i = 12.3 * (di / X_i) ** 2 * (n / 50) ** 2
            else:
                H_i = 12.3 * (di / X_i) ** 2 * (n / 50) ** 2 * 0.95
            Num = Num + 1
            H_d = 0
            SumH = SumH + H_i
            matr.append([Num, round(di, 3), round(li, 3), round(tetta_i, 2), round(rho_i, 3), round(X_i, 3), round(H_i, 2),round(H_d, 2)])
        H_m = SumH / Z
        q_t = 4.8 * 10 ** (-4) * (1 - etaoi) * H0 * (Z - 1) / Z
        Z_new = round(H0 * (1 + q_t) / H_m)
        DeltaZ = abs(Z - Z_new)
        #print(ite, Z)
        Z = Z_new
        ite += 1
    DeltaH = (H0 * (1 + q_t) - SumH) / Z
    a = 0
    for elem in matr:
        matr[a][7] = round(elem[6]+DeltaH,2)
        a += 1


    ## Добавлено для таблицы
    N_=[]
    di_=[]
    li_=[]
    tettai_=[]
    rhoi_=[]
    Xi_=[]
    Hi_=[]
    Hdi_=[]
    a = 0
    for elem in matr:
        N_.append(matr[a][0])
        di_.append(matr[a][1])
        li_.append(matr[a][2])
        tettai_.append(matr[a][3])
        rhoi_.append(matr[a][4])
        Xi_.append(matr[a][5])
        Hi_.append(matr[a][6])
        Hdi_.append(matr[a][7])
        a += 1

    di_ = [float(x) for x in di_]
    li_ = [float(x) for x in li_]
    tettai_ = [float(x) for x in tettai_]
    rhoi_ = [float(x) for x in rhoi_]
    Xi_ = [float(x) for x in Xi_]
    Hi_ = [float(x) for x in Hi_]
    Hdi_ = [float(x) for x in Hdi_]

    ## Таблица
    table=pd.DataFrame( {"№ ступени": (N_),
                           "di, м": (di_),
                           "li, м": (li_),
                           "θi ": (tettai_),
                           "ρi ": (rhoi_),
                           "Xi ": (Xi_),
                           "Hi, кДж/кг": (Hi_),
                           "Hi + Δ, кДж/кг": (Hdi_)
                           }
                       )

    st.dataframe(table)
    st.session_state.table = table

    ## Графики
    z =[]
    for a in range(1, Z+1):
        z.append(a)

    st.write("#")
    fig = plt.figure(figsize=(10, 5))
    ax = fig.gca()
    ax.set_xticks(np.arange(1, z[-1]+1, 1))
    plt.grid(True)
    plt.plot(z, di_, '-ro')
    plt.title('Рисунок 1. Распределение средних диаметров по проточной части')
    plt.savefig('Рисунок 1 Распределение средних диаметров по проточной части.png')
    st.pyplot(fig)

    st.write("#")
    fig = plt.figure(figsize=(10, 5))
    ax = fig.gca()
    ax.set_xticks(np.arange(1, z[-1]+1, 1))
    plt.grid(True)
    plt.plot(z, li_, '-ro')
    plt.title('Рисунок 2. Распределение высот лопаток по проточной части')
    plt.savefig('Рисунок 2 Распределение высот лопаток по проточной части.png')
    st.pyplot(fig)

    st.write("#")
    fig = plt.figure(figsize=(10, 5))
    ax = fig.gca()
    ax.set_xticks(np.arange(1, z[-1]+1, 1))
    plt.grid(True)
    plt.plot(z, tettai_, '-ro')
    plt.title('Рисунок 3. Распределение обратной веерности по проточной части')
    plt.savefig('Рисунок 3 Распределение обратной веерности по проточной части.png')
    st.pyplot(fig)

    st.write("#")
    fig = plt.figure(figsize=(10, 5))
    ax = fig.gca()
    ax.set_xticks(np.arange(1, z[-1]+1, 1))
    plt.grid(True)
    plt.plot(z, rhoi_, '-ro')
    plt.title('Рисунок 4. Распределение степени реактивности по проточной части')
    plt.savefig('Рисунок 4 Распределение степени реактивности по проточной части.png')
    st.pyplot(fig)

    st.write("#")
    fig = plt.figure(figsize=(10, 5))
    ax = fig.gca()
    ax.set_xticks(np.arange(1, z[-1]+1, 1))
    plt.grid(True)
    plt.plot(z, Xi_, '-ro')
    plt.title('Рисунок 5. Распределение U/Cф по проточной части')
    plt.savefig('Рисунок 5 Распределение U_Cф по проточной части.png')
    st.pyplot(fig)

    st.write("#")
    fig = plt.figure(figsize=(10, 5))
    ax = fig.gca()
    ax.set_xticks(np.arange(1, z[-1]+1, 1))
    plt.grid(True)
    plt.plot(z, Hi_, '-ro')
    plt.title('Рисунок 6. Распределение теплоперепадов по проточной части')
    plt.savefig('Рисунок 6 Распределение теплоперепадов по проточной части.png')
    st.pyplot(fig)

    st.write("#")
    fig = plt.figure(figsize=(10, 5))
    ax = fig.gca()
    ax.set_xticks(np.arange(1, z[-1]+1, 1))
    plt.grid(True)
    plt.plot(z, Hdi_, '-ro')
    plt.title('Рисунок 7. Распределение теплоперепадов с учетом невязки по проточной части')
    plt.savefig('Рисунок 7 Распределение теплоперепадов с учетом невязки по проточной части.png')
    st.pyplot(fig)

if page == "Скачать результаты":
    print_xlsx()
    with open("Результаты "+ sheet[st.session_state.index_row][0].value + ".xlsx", "rb") as file:
        st.download_button(
            label="Скачать результаты",
            data=file,
            file_name="Результаты "+ sheet[st.session_state.index_row][0].value + ".xlsx",
            mime='text/xlsx',
        )
    clear()









